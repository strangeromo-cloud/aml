"""Shared helpers for all list-source fetchers.

Each fetcher should:
  - Receive a logger and writes its records via write_excel(...)
  - Raise FetchError with a meaningful message on failure
  - Be tolerant of network blips (use http_get, which retries)
  - Validate at least one row was extracted before declaring success
"""
from __future__ import annotations

import json
import logging
import os
import time
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = REPO_ROOT / "data" / "downloads"
MANIFEST_FILE = OUTPUT_DIR / "_manifest.json"

DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "*/*",
}


class FetchError(RuntimeError):
    """Raised when a fetcher cannot fulfill its contract (network, parse, etc.)."""


def http_get(
    url: str,
    *,
    headers: dict[str, str] | None = None,
    timeout: int = 60,
    retries: int = 3,
    backoff: float = 2.0,
) -> bytes:
    """GET a URL with retries and a sane User-Agent. Returns raw bytes.

    Raises FetchError on persistent failure.
    """
    merged = {**DEFAULT_HEADERS, **(headers or {})}
    last_err: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(url, headers=merged)
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read()
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            last_err = e
            if attempt < retries:
                time.sleep(backoff ** attempt)
            else:
                break
    raise FetchError(f"GET {url} failed after {retries} attempts: {last_err}")


# ---------- Excel writer ----------

ARIAL = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_excel(
    out_path: Path,
    *,
    title: str,
    source_url: str,
    fetched_at: str,
    rows: list[dict[str, Any]],
    columns: list[tuple[str, str, int]],
    extra_about_rows: Iterable[tuple[str, str]] = (),
    sheet_name: str = "Data",
) -> None:
    """Write a 2-sheet workbook (About + Data) for a single source.

    `columns` is a list of (key, header, width).
    """
    wb = Workbook()

    info = wb.active
    info.title = "About"
    info["A1"] = title
    info["A1"].font = Font(name=ARIAL, bold=True, size=16)
    info.merge_cells("A1:C1")
    info["A3"] = "Source URL"
    info["B3"] = source_url
    info["B3"].hyperlink = source_url
    info["B3"].font = Font(name=ARIAL, color="0563C1", underline="single")
    info["A4"] = "Fetched at (UTC)"
    info["B4"] = fetched_at
    info["A5"] = "Records"
    info["B5"] = len(rows)
    info["B5"].font = Font(name=ARIAL, bold=True)
    row_no = 7
    for k, v in extra_about_rows:
        info.cell(row=row_no, column=1, value=k).font = Font(name=ARIAL, bold=True, size=10)
        info.cell(row=row_no, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        row_no += 1
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 90

    data = wb.create_sheet(sheet_name)
    for i, (_, header, width) in enumerate(columns, 1):
        c = data.cell(row=1, column=i, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        data.column_dimensions[get_column_letter(i)].width = width

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _, _) in enumerate(columns, start=1):
            val = row.get(key, "")
            cell = data.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
            cell.font = Font(name=ARIAL, size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER

    if rows:
        data.freeze_panes = "A2"
        data.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------- Manifest ----------

def load_manifest() -> dict[str, Any]:
    if not MANIFEST_FILE.exists():
        return {"updatedAt": None, "sources": {}}
    try:
        return json.loads(MANIFEST_FILE.read_text())
    except json.JSONDecodeError:
        return {"updatedAt": None, "sources": {}}


def save_manifest(manifest: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_FILE.write_text(json.dumps(manifest, indent=2, ensure_ascii=False))


# ---------- Fetcher contract ----------

class Fetcher:
    """Base class for all fetchers. Subclasses must override fetch()."""

    # Stable id used in manifest + output filename
    id: str = ""
    # Display name
    name: str = ""
    # Upstream URL (for logging + Excel about page)
    url: str = ""
    # Output filename in data/downloads/
    out_filename: str = ""
    # If True, the source is JS-rendered or behind auth and needs a headless
    # browser / API key to fetch reliably. The default orchestrator skips these
    # so the daily run doesn't produce repeat alerts on a known limitation.
    requires_headless: bool = False

    def fetch(self) -> list[dict[str, Any]]:
        """Return a list of plain-dict records. Raise FetchError on failure."""
        raise NotImplementedError

    def columns(self) -> list[tuple[str, str, int]]:
        """Override: list of (key, header, width)."""
        raise NotImplementedError

    def extra_about_rows(self) -> list[tuple[str, str]]:
        """Override: any extra info to include on the About sheet."""
        return []


def run_fetcher(
    fetcher: Fetcher,
    logger: logging.Logger,
) -> dict[str, Any]:
    """Run a fetcher, write its Excel, and return a manifest entry."""
    started = datetime.now(timezone.utc)
    out_path = OUTPUT_DIR / fetcher.out_filename
    entry: dict[str, Any] = {
        "name": fetcher.name,
        "url": fetcher.url,
        "outputFile": str(out_path.relative_to(REPO_ROOT)),
        "startedAt": started.isoformat(timespec="seconds"),
    }
    try:
        rows = fetcher.fetch()
        if not rows:
            raise FetchError("fetcher returned 0 rows")
        write_excel(
            out_path,
            title=fetcher.name,
            source_url=fetcher.url,
            fetched_at=started.isoformat(timespec="seconds"),
            rows=rows,
            columns=fetcher.columns(),
            extra_about_rows=fetcher.extra_about_rows(),
        )
        entry.update({
            "status": "success",
            "records": len(rows),
            "finishedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        })
        logger.info(f"  ✓ {fetcher.id} — {len(rows)} records → {out_path.name}")
    except Exception as e:
        entry.update({
            "status": "error",
            "error": f"{type(e).__name__}: {e}",
            "finishedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        })
        logger.error(f"  ✗ {fetcher.id} — {type(e).__name__}: {e}")
    return entry


def setup_logger() -> logging.Logger:
    logger = logging.getLogger("data-refresh")
    if not logger.handlers:
        h = logging.StreamHandler()
        h.setFormatter(logging.Formatter("%(asctime)s  %(message)s", datefmt="%H:%M:%S"))
        logger.addHandler(h)
        logger.setLevel(logging.INFO)
    return logger
