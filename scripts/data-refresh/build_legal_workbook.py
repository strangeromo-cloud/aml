"""Build the Legal-facing 3-sheet workbook, byte-for-byte in the same shape as
`公开名单-CPI-离岸-FATF.xlsx`.

That reference file is the format Legal already reads, so the emailed attachment
reproduces it exactly rather than shipping our three internal workbooks:

  Sheet 1  "CPI <year>"                 A1:D1 merged source note · 排名/国家/分数/阈值标记
  Sheet 2  "Offshore Centres 离岸中心"   A1:B1 merged source note · #/辖区
  Sheet 3  "FATF 黑灰名单"               A1:C1 merged source note · 名单/辖区/含义

Styling matched to the reference: note row Arial 8.5 #595959 on #EFEFEF, header
row Arial 10 bold white on #1F4E5F with a thin bottom rule, body Arial 9.5
centred, freeze at A3, and CPI rows below the threshold tinted #FDECEC.

Data comes from the row snapshots that `refresh_all` writes. When a source could
not be fetched the sheet is still produced, and its note row says so outright
along with where the rows actually came from — never presented as a fresh fetch.
"""
from __future__ import annotations

import json
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = REPO_ROOT / "public" / "downloads"
SNAPSHOT_DIR = OUTPUT_DIR / "_snapshots"
MANIFEST_FILE = OUTPUT_DIR / "_manifest.json"
SEED_DIR = Path(__file__).resolve().parent / "seeds"
# Legal-confirmed overrides written by the upload page. Precedence is by RECENCY, not
# by who produced it: whichever of the two is newer wins, so a confirmed version cannot
# freeze the list and a routine re-fetch cannot discard a deliberate correction.
#   CPI      — compare edition years; a newer TI edition wins, the same edition means
#              Legal corrected it and theirs wins.
#   Offshore — the list carries no version, so compare when the upstream content last
#              actually CHANGED (manifest contentChangedAt) against confirmedAt. Using
#              the fetch time instead would defeat every override, since that is always
#              now.
CPI_OVERRIDE = SEED_DIR / "cpi-override.json"
OFFSHORE_OVERRIDE = SEED_DIR / "offshore-override.json"


def _cpi_override_wins(ov: dict[str, Any], fetched_year: str, *,
                       fetch_usable: bool) -> bool:
    """Who wins for CPI — the confirmed override, or this run's fetch?

    CPI is the one source where "is the fetch newer?" is answerable every run rather
    than by timestamp: the fetcher reads Transparency International's own
    CPI<year>_Results.xlsx after resolving the current edition page. So a usable
    fetch IS the published edition, and Legal's override is a transcription of that
    same file. A same-edition disagreement is therefore a transcription slip, not
    newer data — as Denmark 90-vs-89 turned out to be.

    The override wins when it is genuinely ahead or when we have nothing to compare:
    a newer edition than we could fetch, an unusable fetch, or an uncomparable year.
    """
    ov_year = str(ov.get("edition") or "")
    if not fetch_usable:
        return True          # nothing fetched this run — the override is the fallback
    if not ov_year.isdigit() or not str(fetched_year).isdigit():
        return True          # cannot compare editions — keep the human's version
    if int(ov_year) > int(fetched_year):
        return True          # Legal has an edition we cannot fetch yet
    if int(fetched_year) > int(ov_year):
        print(f"::notice::抓取到更新的 CPI 版本 {fetched_year}（法务确认版本为 {ov_year}），"
              f"以抓取为准")
        return False
    print(f"::notice::CPI 版本相同（{fetched_year}），本次抓取直接读自 TI 发布文件，"
          f"以抓取为准；法务确认版本仅用于比对")
    return False


def _override_newer_than_content(ov: dict[str, Any], entry: dict[str, Any]) -> bool:
    """True when the confirmation is later than the upstream's last real change."""
    changed_at = entry.get("contentChangedAt")
    confirmed = str(ov.get("confirmedAt") or "").strip()
    if not changed_at or not confirmed:
        return True
    # confirmedAt is Beijing "YYYY-MM-DD HH:MM"; contentChangedAt is UTC ISO.
    try:
        c = datetime.strptime(confirmed[:16], "%Y-%m-%d %H:%M").replace(tzinfo=TZ_SHANGHAI)
        u = datetime.fromisoformat(str(changed_at))
    except ValueError:
        return True
    if u > c:
        print(f"::notice::上游离岸名单在法务确认（{confirmed}）之后发生过变化"
              f"（{changed_at}），以抓取为准")
        return False
    return True


def _load_override(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text())
    except json.JSONDecodeError:
        print(f"::warning::{path.name} 无法解析，忽略该覆盖，改用抓取数据")
        return None

TZ_SHANGHAI = timezone(timedelta(hours=8))

# Legal-confirmed threshold from Payment Monitoring Scenarios-0803, Updated-BSR:
# "Check vendor registration country, invoice-issuing country, or receiving-bank
# country's CPI <= 40". Inclusive, and 40 rather than the 31 the earlier reference
# workbook carried as "建议阈值 31（待法务确认）". Matches CPI_THRESHOLD in the demo.
CPI_THRESHOLD = 40

ARIAL = "Arial"
NOTE_FONT = Font(name=ARIAL, size=8.5, color="595959")
NOTE_FILL = PatternFill("solid", fgColor="EFEFEF")
HEAD_FONT = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
BODY_FONT = Font(name=ARIAL, size=9.5)
RULE = Border(bottom=Side(style="thin", color="D9D9D9"))
HI_FILL = PatternFill("solid", fgColor="FDECEC")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_BODY = Alignment(horizontal="center")
LEFT_BODY = Alignment()
# FATF rows are banded by list in the reference file: black = dark row with white
# text, grey = light grey row with default text.
FATF_BLACK_FILL = PatternFill("solid", fgColor="3B3B3B")
FATF_BLACK_FONT = Font(name=ARIAL, size=9.5, color="FFFFFF")
FATF_GREY_FILL = PatternFill("solid", fgColor="D9D9D9")


def _load_snapshot(source_id: str) -> list[dict[str, Any]] | None:
    p = SNAPSHOT_DIR / f"{source_id}.json"
    if not p.exists():
        return None
    try:
        rows = json.loads(p.read_text())
        return rows or None
    except json.JSONDecodeError:
        return None


def _manifest() -> dict:
    if not MANIFEST_FILE.exists():
        return {"sources": {}}
    try:
        return json.loads(MANIFEST_FILE.read_text())
    except json.JSONDecodeError:
        return {"sources": {}}


def _cpi_year() -> str:
    """TI publishes one workbook per edition; the year is in the resolved URL."""
    xlsx = OUTPUT_DIR / "ti-cpi.xlsx"
    if xlsx.exists():
        try:
            from openpyxl import load_workbook
            about = load_workbook(xlsx)["About"]
            for r in range(1, 12):
                if about.cell(r, 1).value == "Source URL":
                    m = re.search(r"CPI(\d{4})", str(about.cell(r, 2).value or ""))
                    if m:
                        return m.group(1)
        except Exception:
            pass
    return ""


def _write_sheet(
    ws,
    note: str,
    headers: list[tuple[str, int]],
    rows: list[list[Any]],
    highlight: set[int] | None = None,
    centered_cols: tuple[int, ...] = (),
    row_style: dict[int, tuple[PatternFill | None, Font | None]] | None = None,
) -> None:
    """`centered_cols` are 1-based column indexes centred in the body, matching the
    reference file (which left-aligns name and note columns). `row_style` applies a
    per-row fill/font override, keyed by 0-based data-row index."""
    ncol = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(1, 1, note)
    c.font = NOTE_FONT
    for i in range(1, ncol + 1):
        ws.cell(1, i).fill = NOTE_FILL

    for i, (h, w) in enumerate(headers, 1):
        hc = ws.cell(2, i, h)
        hc.font = HEAD_FONT
        hc.fill = HEAD_FILL
        hc.alignment = CENTER
        hc.border = RULE
        ws.column_dimensions[get_column_letter(i)].width = w

    highlight = highlight or set()
    row_style = row_style or {}
    for r_i, row in enumerate(rows, start=3):
        idx = r_i - 3
        fill, font = row_style.get(idx, (None, None))
        for c_i in range(1, ncol + 1):
            val = row[c_i - 1] if c_i - 1 < len(row) else None
            cell = ws.cell(r_i, c_i, val)
            cell.font = font or BODY_FONT
            cell.alignment = CENTER_BODY if c_i in centered_cols else LEFT_BODY
            cell.border = RULE
            if idx in highlight:
                cell.fill = HI_FILL
            elif fill is not None:
                cell.fill = fill
    ws.freeze_panes = "A3"


def build(out_path: Path) -> dict:
    """Write the workbook and return a per-sheet report for the email body."""
    man = _manifest()
    src = man.get("sources", {})
    fetched_bj = datetime.now(TZ_SHANGHAI).strftime("%Y-%m-%d")
    report: dict[str, dict] = {}
    wb = Workbook()

    # ── Sheet 1: CPI ────────────────────────────────────────────────────
    year = _cpi_year()
    ws = wb.active
    ws.title = f"CPI {year}".strip()
    cpi = _load_snapshot("ti-cpi")
    entry = src.get("ti-cpi") or {}
    ov = _load_override(CPI_OVERRIDE)
    cpi_fetched = bool(cpi)
    cpi_source = "抓取"
    cpi_reason = "直读 TI 发布文件"
    if ov and ov.get("scores") and _cpi_override_wins(ov, year, fetch_usable=bool(cpi)):
        # Rebuild rows from the confirmed scores, ranking by score so the sheet still
        # reads like the fetched one.
        rows_ov = sorted(({"country": k, "score": v} for k, v in ov["scores"].items()),
                         key=lambda r: (-(r["score"] or 0), r["country"]))
        for i, r in enumerate(rows_ov, 1):
            r["rank"] = i
        cpi, cpi_source = rows_ov, f"法务确认（{ov.get('confirmedBy')} · {ov.get('confirmedAt')}）"
        cpi_reason = ("本次抓取不可用，回退到法务确认版本" if not cpi_fetched
                      else "法务确认版本的版次更新")
        year = ov.get("edition") or year
    if cpi:
        cpi = sorted(cpi, key=lambda r: (r.get("rank") or 9999, str(r.get("country") or "")))
        rows, hi = [], set()
        for i, r in enumerate(cpi):
            score = r.get("score")
            try:
                score_num = int(round(float(score)))
            except (TypeError, ValueError):
                score_num = score
            flagged = isinstance(score_num, int) and score_num <= CPI_THRESHOLD
            rows.append([r.get("rank"), r.get("country"), score_num,
                         "高风险 High-risk" if flagged else None])
            if flagged:
                hi.add(i)
        note = (f"来源: Transparency International — Corruption Perceptions Index {year} "
                f"(https://www.transparency.org/en/cpi/{year}) · 0-100，分数越低越腐败 · "
                f"红底 = CPI ≤ {CPI_THRESHOLD} 判为高风险（法务确认口径，0803 需求表） · "
                f"数据来源: {cpi_source} · {fetched_bj}")
        report["cpi"] = {"ok": True, "records": len(rows), "flagged": len(hi),
                         "year": year, "source": cpi_source, "reason": cpi_reason}
    else:
        rows, hi = [["—", "抓取失败，无数据", "—", "—"]], set()
        note = (f"来源: Transparency International CPI · ⚠ 本次抓取失败"
                f"（{entry.get('error', '未知错误')}），无可用数据 · 抓取尝试: {fetched_bj}")
        report["cpi"] = {"ok": False, "records": 0, "error": entry.get("error"), "year": year}
    _write_sheet(ws, note,
                 [("排名 Rank", 10), ("国家/地区 Country", 32), ("分数 Score", 10),
                  (f"≤/> 阈值{CPI_THRESHOLD}", 16)],
                 rows, hi, centered_cols=(1, 3))

    # ── Sheet 2: Offshore ───────────────────────────────────────────────
    ws = wb.create_sheet("Offshore Centres 离岸中心")
    off = _load_snapshot("eu-offshore-centres")
    entry = src.get("eu-offshore-centres") or {}
    ov = _load_override(OFFSHORE_OVERRIDE)
    off_source = "抓取"
    if ov and ov.get("jurisdictions") and _override_newer_than_content(ov, entry):
        off = [{"jurisdiction": n} for n in ov["jurisdictions"]]
        off_source = f"法务确认（{ov.get('confirmedBy')} · {ov.get('confirmedAt')}）"
    if off:
        names = sorted({str(r.get("jurisdiction") or "").strip() for r in off} - {""})
        rows = [[i, n] for i, n in enumerate(names, 1)]
        note = ("来源: Eurostat Glossary — List of offshore financial centres "
                f"(Balance of Payments Vademecum, Appendix 7) · 共 {len(names)} 个辖区 · "
                f"数据来源: {off_source} · {fetched_bj}")
        report["offshore"] = {"ok": True, "records": len(rows), "source": off_source}
    else:
        rows = [["—", "抓取失败，无数据"]]
        note = ("来源: Eurostat Glossary — List of offshore financial centres · ⚠ 本次抓取失败"
                f"（{entry.get('error', '未知错误')}） · 抓取尝试: {fetched_bj}")
        report["offshore"] = {"ok": False, "records": 0, "error": entry.get("error")}
    _write_sheet(ws, note, [("#", 6), ("辖区 Jurisdiction", 36)], rows,
                 centered_cols=(1,))

    # ── Sheet 3: FATF ───────────────────────────────────────────────────
    def _same_lists(a, b, classify) -> bool:
        """Do two row sets carry the same jurisdictions in the same buckets?"""
        try:
            from .verify_fatf import norm
        except Exception:
            def norm(x):  # type: ignore[misc]
                return str(x or "").strip().lower()

        def key(rows):
            out: dict[str, set[str]] = {}
            for r in rows or []:
                out.setdefault(classify(str(r.get("status") or "")), set()).add(
                    norm(r.get("country")))
            return {k: v - {""} for k, v in out.items()}
        return key(a) == key(b)

    ws = wb.create_sheet("FATF 黑灰名单")
    fatf = _load_snapshot("fatf-jurisdictions")
    entry = src.get("fatf-jurisdictions") or {}
    # The fetcher matches country names across the whole statement page, so a
    # jurisdiction merely *mentioned* (typically one being removed) can land in the
    # wrong list. Validate before this reaches Legal — an implausible list in the
    # attachment is far worse than falling back to the signed-off baseline.
    fetch_rejected: str | None = None
    if fatf:
        try:
            from .verify_fatf import validate_rows
            fetch_rejected = validate_rows(fatf)
        except Exception:
            fetch_rejected = None
        if fetch_rejected:
            fatf = None

    # The Legal-maintained baseline is two things at once: the fallback when the
    # fetch fails, and the authority when it is *newer* than what the fetch
    # returned — the same recency rule the CPI/offshore overrides follow. Without
    # the date comparison a successful-but-stale fetch would silently roll Legal's
    # newer list back.
    try:
        from .verify_fatf import load_baseline, parse_date
        seed = load_baseline() or {}
    except Exception:
        seed = {}

        def parse_date(_):  # type: ignore[misc]
            return None
    seed_rows = seed.get("rows") or None
    seed_date = str(seed.get("listDate") or "")

    fetch_date = ""
    for r in fatf or []:
        d = parse_date(r.get("publication_date"))
        if d and d > fetch_date:
            fetch_date = d

    # Compare at plenary granularity (YYYY-MM): FATF publishes in Feb/Jun/Oct, the
    # baseline carries the statement day while the page often only says "June 2026",
    # so day-level comparison would call a same-plenary fetch "older" every time.
    # The fetch takes over only on a strictly newer plenary — which is exactly the
    # case Legal has not confirmed yet. Within the same plenary the human-confirmed
    # list wins; when the two disagree there, verify_fatf has already alerted.
    fetched_rows = fatf
    prefer_baseline = ""      # "" = use the fetch; otherwise why the baseline won
    if fatf and seed_rows and seed_date and fetch_date:
        if seed_date[:7] > fetch_date[:7]:
            prefer_baseline = "older"
        elif seed_date[:7] == fetch_date[:7]:
            prefer_baseline = "same"
        if prefer_baseline:
            fatf = None

    BLACK = "黑名单 Black"
    GREY = "灰名单 Grey"
    MEANING = {
        BLACK: "严重战略缺陷：强化尽调，最严重者采取反制措施",
        GREY: "在 FATF 监督下整改中：加强监控",
    }

    def _classify(status: str) -> str:
        s = (status or "").lower()
        if "call for action" in s or "black" in s:
            return BLACK
        return GREY

    seeded = False
    list_date = ""
    if not fatf:
        # Falls through to the baseline, whether the fetch failed outright, was
        # rejected as implausible, or came back older than Legal's list. Labelled
        # as the baseline so nobody reads it as a fresh fetch.
        fatf = seed_rows
        list_date = seed_date
        seeded = bool(fatf)

    if fatf:
        buckets: dict[str, list[str]] = {BLACK: [], GREY: []}
        for r in fatf:
            buckets[_classify(str(r.get("status") or ""))].append(
                str(r.get("country") or "").strip())
            if not list_date and r.get("publication_date"):
                list_date = str(r["publication_date"])
        rows = []
        for key in (BLACK, GREY):
            for n in sorted(set(buckets[key]) - {""}):
                rows.append([key, n, MEANING[key]])
        base = ("来源: FATF — High-Risk Jurisdictions subject to a Call for Action(黑) / "
                "Jurisdictions under Increased Monitoring(灰)")
        if seeded:
            agrees = None
            if prefer_baseline == "same":
                agrees = _same_lists(fetched_rows, seed_rows, _classify)
                why = ("与本次官方抓取一致（同一期 " + fetch_date[:7] + "），采用法务确认口径"
                       if agrees else
                       "与本次官方抓取不一致（同一期 " + fetch_date[:7] +
                       "，已触发核验告警），按法务确认口径出表")
            elif prefer_baseline == "older":
                why = (f"本次抓取到的是更早一期（{fetch_date[:7]}），"
                       f"按「以新为准」采用法务基线")
            elif fetch_rejected:
                why = f"本次抓取结果不可信已忽略（{fetch_rejected}）"
            else:
                why = f"本次抓取失败（{entry.get('error', 'Cloudflare 拦截')}）"
            mark = "·" if agrees else "· ⚠"
            note = (f"{base} {mark} {why}，"
                    f"名单日期 {list_date}（法务维护的基线） · 抓取尝试: {fetched_bj}")
            report["fatf"] = {
                "ok": bool(agrees), "seeded": True, "records": len(rows),
                "listDate": list_date, "source": "baseline",
                "preferBaseline": prefer_baseline or None, "agreesWithFetch": agrees,
                "fetchDate": fetch_date or None, "reason": why,
                "error": fetch_rejected or entry.get("error")}
        else:
            note = (f"{base} · 名单日期: {list_date}（每年 2/6/10 月更新，需定期刷新） · "
                    f"抓取: {fetched_bj}")
            report["fatf"] = {"ok": True, "seeded": False, "records": len(rows),
                              "listDate": list_date, "source": "fetch",
                              "reason": "本次官方抓取，且是比法务基线更新的一期"
                                        if seed_date else "本次官方抓取"}
    else:
        rows = [["—", "抓取失败，无数据", "—"]]
        note = (f"来源: FATF 黑/灰名单 · ⚠ 本次抓取失败（{entry.get('error', '未知错误')}），"
                f"且无历史基线 · 抓取尝试: {fetched_bj}")
        report["fatf"] = {"ok": False, "seeded": False, "records": 0,
                          "error": entry.get("error")}
    fatf_style = {
        i: ((FATF_BLACK_FILL, FATF_BLACK_FONT) if r[0] == BLACK else (FATF_GREY_FILL, None))
        for i, r in enumerate(rows) if r[0] in (BLACK, GREY)
    }
    _write_sheet(ws, note,
                 [("名单 List", 14), ("辖区 Jurisdiction", 34), ("含义", 44)], rows,
                 row_style=fatf_style)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    report["_file"] = {"path": str(out_path), "sheets": wb.sheetnames}
    return report


def default_out_path() -> Path:
    stamp = datetime.now(TZ_SHANGHAI).strftime("%Y%m%d")
    return OUTPUT_DIR / f"公开名单-CPI-离岸-FATF-{stamp}.xlsx"


if __name__ == "__main__":
    p = default_out_path()
    rep = build(p)
    print(json.dumps(rep, ensure_ascii=False, indent=1))
