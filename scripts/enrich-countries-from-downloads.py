"""Merge real WJP Rule of Law + TJN Financial Secrecy scores from the daily
downloads into data/countries.json.

Run after a daily refresh, or any time you want to upgrade countries.json
with the latest values from public/downloads/wjp-rule-of-law.xlsx and
public/downloads/tjn-fsi.xlsx.

Adds two optional fields to each country:
  wjpRoLScore   :  WJP Rule of Law Index, 0..1 (higher = stronger rule of law)
  tjnSecrecyScore : Tax Justice Network FSI secrecy score, 0..100 (higher = more secret)

If a country isn't in the upstream dataset, no field is added — the scoring
code falls back to its proxy formula.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
COUNTRIES_JSON = ROOT / "data" / "countries.json"
WJP_XLSX = ROOT / "public" / "downloads" / "wjp-rule-of-law.xlsx"
TJN_XLSX = ROOT / "public" / "downloads" / "tjn-fsi.xlsx"

# Minimal ISO-2 → ISO-3 mapping for the 55 countries we ship in countries.json.
# Generated once; if you add a new country to countries.json, add its ISO-3 here.
ISO2_TO_ISO3 = {
    "US": "USA", "CA": "CAN", "MX": "MEX", "BR": "BRA", "AR": "ARG",
    "CL": "CHL", "CO": "COL", "PE": "PER", "VE": "VEN", "CU": "CUB",
    "HT": "HTI",
    "GB": "GBR", "DE": "DEU", "FR": "FRA", "IT": "ITA", "ES": "ESP",
    "NL": "NLD", "BE": "BEL", "CH": "CHE", "SE": "SWE", "FI": "FIN",
    "DK": "DNK", "NO": "NOR", "IE": "IRL", "AT": "AUT", "PT": "PRT",
    "PL": "POL", "CZ": "CZE", "HU": "HUN", "RO": "ROU", "BG": "BGR",
    "HR": "HRV", "LU": "LUX", "MC": "MCO",
    "RU": "RUS", "BY": "BLR", "UA": "UKR", "TR": "TUR",
    "CN": "CHN", "HK": "HKG", "TW": "TWN", "JP": "JPN", "KR": "KOR",
    "KP": "PRK", "SG": "SGP", "MY": "MYS", "TH": "THA", "VN": "VNM",
    "ID": "IDN", "PH": "PHL", "IN": "IND", "PK": "PAK", "BD": "BGD",
    "MM": "MMR", "AU": "AUS", "NZ": "NZL",
    "AE": "ARE", "SA": "SAU", "IL": "ISR", "IR": "IRN", "SY": "SYR",
    "YE": "YEM", "LB": "LBN",
    "EG": "EGY", "ZA": "ZAF", "NG": "NGA", "KE": "KEN", "ML": "MLI", "CD": "COD",
}


def load_wjp_scores() -> dict[str, float]:
    """Return ISO-3 → overall_score (0..1)."""
    if not WJP_XLSX.exists():
        print(f"  ! {WJP_XLSX} missing; WJP scores will be empty")
        return {}
    wb = load_workbook(WJP_XLSX, data_only=True, read_only=True)
    ws = wb["Data"]
    out: dict[str, float] = {}
    for r in range(2, ws.max_row + 1):
        iso3 = ws.cell(row=r, column=2).value  # "ISO Code" column
        score = ws.cell(row=r, column=5).value  # "Overall Score (0-1)" column
        if isinstance(iso3, str) and isinstance(score, (int, float)):
            out[iso3] = float(score)
    return out


def load_tjn_scores() -> dict[str, dict]:
    """Return ISO-2 → {secrecy: 0..100, fsi_value, rank}."""
    if not TJN_XLSX.exists():
        print(f"  ! {TJN_XLSX} missing; TJN scores will be empty")
        return {}
    wb = load_workbook(TJN_XLSX, data_only=True, read_only=True)
    ws = wb["Data"]
    out: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        rank = ws.cell(row=r, column=1).value
        country = ws.cell(row=r, column=2).value
        iso2 = ws.cell(row=r, column=3).value
        fsi_value = ws.cell(row=r, column=4).value
        secrecy = ws.cell(row=r, column=5).value
        if isinstance(iso2, str) and isinstance(secrecy, (int, float)):
            out[iso2] = {
                "rank": rank,
                "name": country,
                "secrecy_score": float(secrecy),
                "fsi_value": float(fsi_value) if isinstance(fsi_value, (int, float)) else None,
            }
    return out


def main() -> int:
    countries = json.loads(COUNTRIES_JSON.read_text())
    wjp = load_wjp_scores()
    tjn = load_tjn_scores()
    print(f"Loaded {len(wjp)} WJP rows, {len(tjn)} TJN rows")

    wjp_hits = tjn_hits = 0
    misses_wjp: list[str] = []
    misses_tjn: list[str] = []
    for c in countries:
        iso2 = c.get("code")
        iso3 = ISO2_TO_ISO3.get(iso2)
        if iso3 and iso3 in wjp:
            c["wjpRoLScore"] = round(wjp[iso3], 4)
            wjp_hits += 1
        else:
            c.pop("wjpRoLScore", None)
            misses_wjp.append(iso2 or "?")
        if iso2 and iso2 in tjn:
            c["tjnSecrecyScore"] = round(tjn[iso2]["secrecy_score"], 2)
            tjn_hits += 1
        else:
            c.pop("tjnSecrecyScore", None)
            misses_tjn.append(iso2 or "?")

    COUNTRIES_JSON.write_text(json.dumps(countries, indent=2, ensure_ascii=False))
    print(f"\nResult:")
    print(f"  WJP coverage: {wjp_hits}/{len(countries)} (missing: {misses_wjp})")
    print(f"  TJN coverage: {tjn_hits}/{len(countries)} (missing: {misses_tjn})")
    print(f"  Wrote: {COUNTRIES_JSON.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
