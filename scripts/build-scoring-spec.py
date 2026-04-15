"""Generate a multi-sheet Excel describing the AML risk scoring spec."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Style helpers ----------
ARIAL = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F2937")  # slate-800
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
DIM_FILLS = {
    "Sanctions Screening": PatternFill("solid", start_color="FECACA"),       # rose
    "Country Risk Scoring": PatternFill("solid", start_color="FED7AA"),       # orange
    "High-Risk Jurisdiction Monitoring": PatternFill("solid", start_color="FEF3C7"),  # amber
    "Sanctions Circumvention Risk": PatternFill("solid", start_color="DDD6FE"),       # violet
    "PEP & Adverse Media": PatternFill("solid", start_color="BAE6FD"),                # sky
    "Country Context Enrichment": PatternFill("solid", start_color="BBF7D0"),         # emerald
}
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---------- Data ----------
DIMENSIONS = [
    ("Sanctions Screening", "制裁名单筛查", 0.25, "查实体/UBO 是否直接命中各国权威制裁名单"),
    ("Country Risk Scoring", "国家风险评分", 0.20, "仅凭总部所在国推断风险水平"),
    ("High-Risk Jurisdiction Monitoring", "高风险司法辖区监控", 0.15, "动态监控公司业务足迹触碰多少高风险辖区"),
    ("Sanctions Circumvention Risk", "制裁规避风险", 0.15, "识别绕开制裁的金融或物流路径"),
    ("PEP & Adverse Media", "PEP 与负面媒体", 0.15, "受益人政治背景、负面新闻、监管处罚历史"),
    ("Country Context Enrichment", "国家背景增强", 0.10, "治理、法治、金融透明度的国家背景"),
]

# Each row: (dim_en, factor_id, factor_en, factor_zh, weight_within_dim, meaning, rule, sources, how_obtained)
FACTORS = [
    # 1. Sanctions Screening
    ("Sanctions Screening", "sanctions_direct_hit",
     "Direct Sanctions Hit", "直接制裁命中", 0.60,
     "公司总部注册国是否被全面制裁或列入 FATF 黑名单",
     "OFAC 全面 / FATF 黑名单 / UN / EU 制裁 → 100；OFAC 部分制裁 → 75；其他 → 0",
     "OFAC SDN; OFAC Country Programs; UN Consolidated; EU Consolidated; FATF Lists",
     "真实公开数据快照，手工录入到 data/countries.json 的 ofacComprehensive / ofacPartial / unSanctioned / euSanctioned / fatfStatus 字段"),
    ("Sanctions Screening", "sanctions_ubo_hit",
     "UBO Sanctions Exposure", "UBO 制裁暴露", 0.30,
     "最终受益人中有多少居住在受制裁辖区",
     "每位受制裁辖区的 UBO 贡献 60 分，累加封顶 100",
     "OFAC SDN; OpenSanctions",
     "UBO 列表完全 mock — 由 scripts/generate-companies.ts 按公司风险等级随机生成，国家命中判断使用真实国家数据"),
    ("Sanctions Screening", "sanctions_fuzzy_match",
     "Fuzzy Watchlist Match", "模糊名单匹配", 0.10,
     "公司名与综合制裁名单条目的文本相似度",
     "直接把 0-100 相似度作为分数",
     "OpenSanctions; OFAC SDN",
     "fuzzyWatchlistMatchPct 字段在 companies.json 中是 mock 值，按风险等级赋分（Critical 70-99 / High 40-69 / Medium 15-34 / Low 0-9）"),

    # 2. Country Risk Scoring
    ("Country Risk Scoring", "hq_fatf_status",
     "HQ FATF Status", "总部 FATF 状态", 0.40,
     "总部所在国是否在 FATF 高风险或受监控名单",
     "Blacklist → 100；Greylist → 65；Compliant → 10",
     "FATF High-Risk & Monitored Jurisdictions",
     "参照 FATF 2026-02 公开声明手工录入 countries.json 的 fatfStatus 字段"),
    ("Country Risk Scoring", "hq_basel_aml_index",
     "Basel AML Index", "巴塞尔反洗钱指数", 0.30,
     "Basel Institute on Governance 年度反洗钱风险评分（0-10）",
     "线性映射：Basel 值 × 10 → 0-100",
     "Basel AML Index (index.baselgovernance.org)",
     "录入 Basel 2024 报告中各国分数到 countries.json 的 baselAmlIndex 字段（如缅甸 8.17、海地 7.92、芬兰 2.98）"),
    ("Country Risk Scoring", "hq_cpi",
     "Corruption Perceptions Index", "清廉指数", 0.30,
     "Transparency International 清廉指数（0=极腐败，100=极清廉）",
     "反向映射：100 - CPI",
     "Transparency International CPI 2024",
     "录入 TI 2024 报告各国分数到 countries.json 的 cpi 字段（如丹麦 90、中国 43、俄罗斯 22）"),

    # 3. High-Risk Jurisdiction Monitoring
    ("High-Risk Jurisdiction Monitoring", "operating_in_high_risk_pct",
     "High-Risk Operating Footprint", "高风险运营足迹", 0.40,
     "公司运营国家中被列为 FATF 黑/灰名单或 OFAC 制裁的占比",
     "(高风险运营国数 / 总运营国数) × 100",
     "FATF Lists; OFAC Country Programs",
     "operatingCountries 列表 mock（按公司风险等级从不同国家池抽样 4-8 个），运行时用真实国家数据判断高风险"),
    ("High-Risk Jurisdiction Monitoring", "subsidiaries_in_sanctioned_countries",
     "Subsidiaries in Sanctioned Jurisdictions", "受制裁辖区子公司", 0.40,
     "子公司设在 OFAC 全面制裁或 FATF 黑名单国家的数量",
     "0 个 → 0；任一 → 40；每加 1 个 +25；封顶 100",
     "OFAC Country Programs; FATF Lists",
     "subsidiaries 数组完全 mock；Critical 公司 40% 概率把子公司放在 KP/IR/SY，High 35%，等等"),
    ("High-Risk Jurisdiction Monitoring", "sudden_jurisdiction_expansion",
     "Recent High-Risk Expansion", "近期高风险扩张", 0.20,
     "过去 6 个月是否新进入了高风险辖区（动态监控信号）",
     "布尔触发 → 85；否则 → 0",
     "Control Risks Country Risk Forecast",
     "suddenJurisdictionExpansion 布尔字段 mock；Critical 100% / High 60% / Medium 20% / Low 0%"),

    # 4. Sanctions Circumvention Risk
    ("Sanctions Circumvention Risk", "transit_country_pattern",
     "Transit Country Pattern", "中转国家模式", 0.50,
     "公司是否同时在中转枢纽（AE/TR/HK/SG/CY/VG/KY/LU/CH/MC）和受制裁辖区运营",
     "同时命中 → 85；只占其一 → 35；都没有 → 0",
     "Control Risks; OpenSanctions; FinCEN Advisories",
     "TRANSIT_HUBS 集合硬编码在 lib/scoring/circumvention.ts，operatingCountries 仍为 mock"),
    ("Sanctions Circumvention Risk", "neighboring_sanctioned_exposure",
     "Sanctioned-Neighbor Exposure", "制裁国邻国暴露", 0.30,
     "在与全面制裁国接壤的国家做业务的数量（如 KZ/AM/GE 紧邻俄罗斯）",
     "0 个 → 0；任一 → 25；每加 1 个 +20；封顶 100",
     "Control Risks; FinCEN Advisories",
     "SANCTIONED_ADJACENT 集合硬编码在 circumvention.ts，operatingCountries mock"),
    ("Sanctions Circumvention Risk", "opaque_ownership_chain",
     "Opaque Ownership Chain", "不透明所有权链", 0.20,
     "子公司和 UBO 中位于金融保密枢纽的数量（FSI 排名靠前）",
     "子公司数 × 15 + UBO 数 × 25，封顶 100",
     "Tax Justice Network FSI; OpenSanctions",
     "复用 TRANSIT_HUBS 作为保密辖区近似；子公司/UBO 数据 mock"),

    # 5. PEP & Adverse Media
    ("PEP & Adverse Media", "ubo_pep_status",
     "UBO PEP Status", "UBO 政治敏感人物", 0.50,
     "受益人是否为政治敏感人物（现/前任官员、军方、国企高管或直系亲属）",
     "0 位 → 0；1 位 → 45；每加 1 位 +25；封顶 100",
     "OpenSanctions PEP",
     "ubos[].pep 布尔字段 mock；Critical 85% / High 60% / Medium 25% / Low 3% 概率赋 PEP"),
    ("PEP & Adverse Media", "adverse_media_count",
     "Adverse Media Count", "负面媒体数量", 0.30,
     "过去 24 个月涉及反洗钱/欺诈/制裁/腐败等关键词的负面新闻条数",
     "条数 × 8，封顶 100",
     "OpenSanctions; (商业可选: Dow Jones / Refinitiv / GDELT)",
     "adverseMediaCount 数值 mock；Critical 6-15 / High 2-7 / Medium 0-2 / Low 0；可改造为 GDELT API 实时查询"),
    ("PEP & Adverse Media", "regulatory_enforcement_history",
     "Regulatory Enforcement History", "监管处罚历史", 0.20,
     "历史上是否被 OFAC / FinCEN / 欧盟或本国监管机构正式处罚",
     "0 次 → 0；1 次 → 55；每加 1 次 +25；封顶 100",
     "OFAC Enforcement; OpenSanctions",
     "regulatoryEnforcementCount mock；Critical 1-3 次 / High 40% 概率 1 次 / Medium/Low 0"),

    # 6. Country Context Enrichment
    ("Country Context Enrichment", "wgi_control_of_corruption",
     "WGI Control of Corruption", "WGI 腐败控制", 0.50,
     "World Bank 全球治理指标的腐败控制分项（-2.5 到 2.5，越高越好）",
     "标准化：((2.5 - wgi) / 5) × 100",
     "World Bank Worldwide Governance Indicators",
     "WGI 2024 真实分数手工录入 countries.json 的 wgiControlOfCorruption 字段（丹麦 2.31、芬兰 2.17、中国 -0.27、朝鲜 -1.55）"),
    ("Country Context Enrichment", "rule_of_law_index",
     "Rule of Law Environment", "法治环境", 0.30,
     "World Justice Project 法治指数（政府约束、腐败、基本权利、秩序、监管执行）",
     "代理映射：基于 WGI 标准化",
     "WJP Rule of Law Index; (fallback: World Bank WGI)",
     "代理值：lib/scoring/enrichment.ts 的 ruleOfLawProxy() 用 WGI 近似；如要精确应导入 WJP CSV"),
    ("Country Context Enrichment", "financial_transparency",
     "Financial Secrecy", "金融保密度", 0.20,
     "Tax Justice Network 金融保密指数（衡量受益所有权透明度与银行保密程度）",
     "直接使用保密分数 0-100（越高越保密）",
     "Tax Justice Network Financial Secrecy Index",
     "在 enrichment.ts 中维护 hubs 字典手工分档（UAE 78、Monaco 70、新加坡 70、瑞士 65、芬兰 28）"),
]

DATA_SOURCES = [
    ("fatf_lists", "Financial Action Task Force",
     "FATF High-Risk & Monitored Jurisdictions",
     "https://www.fatf-gafi.org/en/publications/High-risk-and-other-monitored-jurisdictions.html",
     "2026-02-14",
     "受呼吁采取行动的黑名单与受加强监控的灰名单"),
    ("ofac_sdn", "U.S. Treasury OFAC",
     "Specially Designated Nationals (SDN) List",
     "https://sanctionssearch.ofac.treas.gov/",
     "2026-03-28",
     "美国受 OFAC 项目资产冻结的个人与实体主要综合名单"),
    ("ofac_countries", "U.S. Treasury OFAC",
     "OFAC Sanctions Programs by Country",
     "https://ofac.treasury.gov/sanctions-programs-and-country-information",
     "2026-03-28",
     "按国家划分的制裁项目说明（古巴/伊朗/朝鲜/叙利亚为全面，俄/委为部门）"),
    ("un_consolidated", "United Nations Security Council",
     "UN Consolidated Sanctions List",
     "https://www.un.org/securitycouncil/content/un-sc-consolidated-list",
     "2026-03-20",
     "受联合国安理会各制裁机制约束的个人与实体清单"),
    ("eu_consolidated", "European Commission",
     "EU Consolidated Financial Sanctions List",
     "https://webgate.ec.europa.eu/fsd/fsf",
     "2026-03-22",
     "受欧盟金融制裁的个人、团体与实体综合清单"),
    ("basel_aml", "Basel Institute on Governance",
     "Basel AML Index",
     "https://index.baselgovernance.org",
     "2024-11-15",
     "基于 17 个指标的国家洗钱与恐怖融资风险年度排名（0-10）"),
    ("ti_cpi", "Transparency International",
     "Corruption Perceptions Index",
     "https://www.transparency.org/en/cpi",
     "2024-01-30",
     "对各国公共部门腐败感知的年度排名（0=极腐败，100=极清廉）"),
    ("wb_wgi", "World Bank Group",
     "Worldwide Governance Indicators",
     "https://info.worldbank.org/governance/wgi/",
     "2024-09-24",
     "六大治理维度（含腐败控制），评分范围 -2.5 到 2.5"),
    ("wjp_rol", "World Justice Project",
     "Rule of Law Index",
     "https://worldjusticeproject.org/rule-of-law-index",
     "2024-10-23",
     "法治评分（政府约束、腐败、基本权利、秩序、监管执行）"),
    ("tjn_fsi", "Tax Justice Network",
     "Financial Secrecy Index",
     "https://taxjustice.net/topics/financial-secrecy-index/",
     "2022-05-17",
     "按各辖区对全球金融保密的贡献度进行排名"),
    ("open_sanctions", "OpenSanctions e.V.",
     "OpenSanctions",
     "https://www.opensanctions.org",
     "2026-03-30",
     "聚合的开放制裁、PEP 及关注名单数据库"),
    ("control_risks_geo", "Control Risks",
     "Country Risk Forecast / RiskMap",
     "https://www.controlrisks.com/our-thinking/insights/riskmap",
     "2025-12-10",
     "商业地缘政治与运营风险评级（用于规避与司法辖区监控）"),
]

RISK_LEVELS = [
    ("Low", "低", "0 – 24", "10b981", "无需特别关注，常规尽职调查"),
    ("Medium", "中", "25 – 49", "f59e0b", "标准 CDD（客户尽职调查）"),
    ("High", "高", "50 – 69", "f97316", "升级至 EDD（加强尽职调查），合规人工复核"),
    ("Critical", "极高", "70 – 100", "ef4444", "拒绝准入或立即上报；触发 SAR 评估"),
]

DATA_REALITY = [
    ("🟢 真实公开数据快照", "FATF 名单状态、OFAC/UN/EU 制裁标记、Basel AML Index、CPI、WGI",
     "data/countries.json 中 55 个国家的 7 个字段",
     "手工从各机构官网最新公开报告录入，无 API 调用"),
    ("🟢 真实数据源 URL 与快照日期", "12 个权威数据源的真实链接",
     "data/sources.json",
     "手工整理，UI 因子卡片中可点击跳转，但本地数据是快照"),
    ("🟡 硬编码分类字典", "中转枢纽清单、制裁邻国清单、金融保密分档、法治代理映射",
     "lib/scoring/*.ts 中的常量（TRANSIT_HUBS / SANCTIONED_ADJACENT / hubs / ruleOfLawProxy）",
     "基于 FinCEN/EU/OFAC 公告与 TJN 排名手工编写"),
    ("🔴 完全 Mock", "200 家公司的名字、行业、运营国、子公司、UBO；adverseMediaCount、PEP、enforcement 等信号",
     "data/companies.json（由 scripts/generate-companies.ts 用种子 PRNG 生成）",
     "按公司风险等级反向设计：先定 Critical/High/Medium/Low，再编出对应的信号"),
    ("🔴 真实公司名（Intel/TSMC 等 30 家）", "公司名为真实但所有评分字段为 mock，强制锁定 Low 风险",
     "REAL_CUSTOMERS 与 REAL_SUPPLIERS 列表",
     "仅用于演示低风险 Tier-1 实体画像；UI 显示 Illustrative 警告"),
]

# ---------- Build workbook ----------
wb = Workbook()

# Sheet 1: Overview
ov = wb.active
ov.title = "评分体系总览"

ov["A1"] = "Lenovo AML Risk Watch — 评分体系完整说明"
ov["A1"].font = Font(name=ARIAL, bold=True, size=16, color="111827")
ov.merge_cells("A1:F1")

ov["A3"] = "总分公式"
ov["A3"].font = Font(name=ARIAL, bold=True, size=12)
ov["A4"] = "总分 = Σ (维度得分 × 维度权重) / Σ 权重，结果在 0–100 区间"
ov["A4"].font = Font(name=ARIAL, size=11)
ov.merge_cells("A4:F4")

# Dimension weights table
ov["A6"] = "维度权重分配"
ov["A6"].font = Font(name=ARIAL, bold=True, size=12)

dim_headers = ["#", "维度（英文）", "维度（中文）", "权重", "回答的核心问题"]
for i, h in enumerate(dim_headers, 1):
    c = ov.cell(row=7, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER

for i, (en, zh, w, q) in enumerate(DIMENSIONS, 1):
    row = 7 + i
    ov.cell(row=row, column=1, value=i).alignment = CENTER
    ov.cell(row=row, column=2, value=en).font = Font(name=ARIAL, size=11)
    ov.cell(row=row, column=3, value=zh).font = Font(name=ARIAL, size=11)
    wc = ov.cell(row=row, column=4, value=w)
    wc.number_format = "0.0%"
    wc.alignment = CENTER
    wc.font = Font(name=ARIAL, bold=True, size=11)
    ov.cell(row=row, column=5, value=q).alignment = WRAP
    for col in range(1, 6):
        ov.cell(row=row, column=col).border = BORDER
        ov.cell(row=row, column=col).fill = DIM_FILLS[en]

# Total row with formula
total_row = 7 + len(DIMENSIONS) + 1
ov.cell(row=total_row, column=3, value="合计").font = Font(name=ARIAL, bold=True, size=11)
total_cell = ov.cell(row=total_row, column=4, value=f"=SUM(D8:D{7+len(DIMENSIONS)})")
total_cell.number_format = "0.0%"
total_cell.font = Font(name=ARIAL, bold=True, size=11)
total_cell.alignment = CENTER
total_cell.border = BORDER

# Risk level table
risk_start = total_row + 3
ov.cell(row=risk_start - 1, column=1, value="风险等级阈值").font = Font(name=ARIAL, bold=True, size=12)
risk_headers = ["等级（英）", "等级（中）", "分数区间", "颜色", "对应行动"]
for i, h in enumerate(risk_headers, 1):
    c = ov.cell(row=risk_start, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER

for i, (en, zh, rng, color, action) in enumerate(RISK_LEVELS, 1):
    r = risk_start + i
    ov.cell(row=r, column=1, value=en).font = Font(name=ARIAL, bold=True, color=color)
    ov.cell(row=r, column=2, value=zh).font = Font(name=ARIAL, bold=True, color=color)
    ov.cell(row=r, column=3, value=rng).alignment = CENTER
    ov.cell(row=r, column=4, value=f"#{color.upper()}").font = Font(name=ARIAL, color=color, bold=True)
    ov.cell(row=r, column=4).alignment = CENTER
    ov.cell(row=r, column=5, value=action).alignment = WRAP
    for col in range(1, 6):
        ov.cell(row=r, column=col).border = BORDER

# Data reality
dr_start = risk_start + len(RISK_LEVELS) + 3
ov.cell(row=dr_start - 1, column=1, value="数据真实度分层").font = Font(name=ARIAL, bold=True, size=12)
dr_headers = ["真实度", "覆盖内容", "在哪", "怎么获得"]
for i, h in enumerate(dr_headers, 1):
    c = ov.cell(row=dr_start, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER

for i, (level, content, where, how) in enumerate(DATA_REALITY, 1):
    r = dr_start + i
    ov.cell(row=r, column=1, value=level).font = Font(name=ARIAL, bold=True, size=11)
    ov.cell(row=r, column=2, value=content).alignment = WRAP
    ov.cell(row=r, column=3, value=where).alignment = WRAP
    ov.cell(row=r, column=4, value=how).alignment = WRAP
    for col in range(1, 5):
        ov.cell(row=r, column=col).border = BORDER

# Column widths
for col, w in zip("ABCDEF", [22, 38, 28, 14, 60, 12]):
    ov.column_dimensions[col].width = w
ov.row_dimensions[1].height = 26

# Sheet 2: Full factor catalog (the headline sheet)
fc = wb.create_sheet("18 因子完整目录")

fc["A1"] = "18 个因子完整说明 — 维度 / 子因子 / 评分逻辑 / 数据来源 / 数据获取方式"
fc["A1"].font = Font(name=ARIAL, bold=True, size=14)
fc.merge_cells("A1:I1")

headers = [
    "#",
    "维度",
    "维度权重",
    "因子 (英)",
    "因子 (中)",
    "因子权重\n(本维度内)",
    "因子含义",
    "评分规则",
    "数据来源",
    "数据获取方式",
]
for i, h in enumerate(headers, 1):
    c = fc.cell(row=3, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER

dim_weight = {en: w for en, _, w, _ in DIMENSIONS}

for i, row in enumerate(FACTORS, 1):
    (dim_en, fid, fen, fzh, w_in, meaning, rule, sources, how) = row
    r = 3 + i
    fc.cell(row=r, column=1, value=i).alignment = CENTER
    fc.cell(row=r, column=2, value=dim_en).font = Font(name=ARIAL, bold=True, size=11)
    wc = fc.cell(row=r, column=3, value=dim_weight[dim_en])
    wc.number_format = "0.0%"
    wc.alignment = CENTER
    fc.cell(row=r, column=4, value=fen)
    fc.cell(row=r, column=5, value=fzh)
    fwc = fc.cell(row=r, column=6, value=w_in)
    fwc.number_format = "0.0%"
    fwc.alignment = CENTER
    fc.cell(row=r, column=7, value=meaning).alignment = WRAP
    fc.cell(row=r, column=8, value=rule).alignment = WRAP
    fc.cell(row=r, column=9, value=sources).alignment = WRAP
    fc.cell(row=r, column=10, value=how).alignment = WRAP
    for col in range(1, 11):
        cell = fc.cell(row=r, column=col)
        cell.border = BORDER
        cell.fill = DIM_FILLS[dim_en]
        if cell.font.size is None:
            cell.font = Font(name=ARIAL, size=10)
        else:
            cell.font = Font(name=ARIAL, bold=cell.font.bold, size=10, color=cell.font.color or "000000")

# Sub-weight check column (verify each dimension's factor weights sum to 100%)
check_row = 3 + len(FACTORS) + 2
fc.cell(row=check_row, column=1, value="校验").font = Font(name=ARIAL, bold=True, size=11)
fc.cell(row=check_row, column=2, value="每个维度因子权重合计应为 100%").font = Font(name=ARIAL, italic=True, size=10)
fc.merge_cells(start_row=check_row, start_column=2, end_row=check_row, end_column=4)

# Build per-dimension SUMIF check rows
dims_seen = []
for r in range(check_row + 1, check_row + 1 + len(DIMENSIONS)):
    idx = r - check_row - 1
    dim_en = DIMENSIONS[idx][0]
    fc.cell(row=r, column=2, value=dim_en).font = Font(name=ARIAL, size=10)
    formula = f'=SUMIF(B4:B{3+len(FACTORS)},"{dim_en}",F4:F{3+len(FACTORS)})'
    sumcell = fc.cell(row=r, column=3, value=formula)
    sumcell.number_format = "0.0%"
    sumcell.alignment = CENTER
    sumcell.font = Font(name=ARIAL, bold=True, size=10)
    sumcell.border = BORDER

# Column widths
widths = [4, 28, 10, 30, 24, 10, 38, 38, 28, 50]
for i, w in enumerate(widths, 1):
    fc.column_dimensions[get_column_letter(i)].width = w
fc.row_dimensions[3].height = 36
for r in range(4, 4 + len(FACTORS)):
    fc.row_dimensions[r].height = 95
fc.row_dimensions[1].height = 22

# Freeze top header
fc.freeze_panes = "C4"

# Sheet 3: Data source catalog
ds = wb.create_sheet("数据源目录")
ds["A1"] = "12 个权威数据源完整目录"
ds["A1"].font = Font(name=ARIAL, bold=True, size=14)
ds.merge_cells("A1:F1")

ds_headers = ["#", "数据源 ID", "权威机构", "完整名称", "URL", "快照日期", "描述"]
for i, h in enumerate(ds_headers, 1):
    c = ds.cell(row=3, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER

for i, (sid, authority, name, url, date, desc) in enumerate(DATA_SOURCES, 1):
    r = 3 + i
    ds.cell(row=r, column=1, value=i).alignment = CENTER
    ds.cell(row=r, column=2, value=sid).font = Font(name="Courier New", size=10)
    ds.cell(row=r, column=3, value=authority).font = Font(name=ARIAL, bold=True, size=10)
    ds.cell(row=r, column=4, value=name)
    link_cell = ds.cell(row=r, column=5, value=url)
    link_cell.hyperlink = url
    link_cell.font = Font(name=ARIAL, color="0563C1", underline="single", size=10)
    ds.cell(row=r, column=6, value=date).alignment = CENTER
    ds.cell(row=r, column=7, value=desc).alignment = WRAP
    for col in range(1, 8):
        cell = ds.cell(row=r, column=col)
        cell.border = BORDER
        if col != 1 and col != 6:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

ds.column_dimensions["A"].width = 4
ds.column_dimensions["B"].width = 18
ds.column_dimensions["C"].width = 30
ds.column_dimensions["D"].width = 35
ds.column_dimensions["E"].width = 55
ds.column_dimensions["F"].width = 13
ds.column_dimensions["G"].width = 50
for r in range(4, 4 + len(DATA_SOURCES)):
    ds.row_dimensions[r].height = 50
ds.freeze_panes = "C4"

# Sheet 4: Hardcoded country sets used in scoring
hs = wb.create_sheet("硬编码国家集合")
hs["A1"] = "评分逻辑中的硬编码国家清单（lib/scoring/circumvention.ts 与 enrichment.ts）"
hs["A1"].font = Font(name=ARIAL, bold=True, size=13)
hs.merge_cells("A1:E1")

hardcoded = [
    ("TRANSIT_HUBS", "中转枢纽", "AE, TR, HK, SG, MC, CY, VG, KY, LU, CH",
     "金融基础设施发达 + 跨境便利 + 历史上被规避路径利用",
     "Sanctions Circumvention Risk → Transit Country Pattern"),
    ("SANCTIONED_ADJACENT", "制裁国邻国", "AM, AZ, KZ, UZ, TJ, TM, GE, IQ, AF, CN, KR, RU, LB, JO",
     "本身不受全面制裁，但地理上紧邻受制裁国，常被用于二次转运",
     "Sanctions Circumvention Risk → Sanctioned-Neighbor Exposure"),
    ("CRITICAL_HQ_ROTATION", "Critical 公司 HQ 轮转",
     "IR, KP, SY, MM, CU",
     "保证生成的 Critical 公司覆盖五个主要受制裁辖区",
     "scripts/generate-companies.ts （生成器）"),
    ("HIGH_RISK_HQ_ROTATION", "High 公司 HQ 轮转",
     "RU, BY, VE, NG, PK, ML, CD, LB, BD, HT, YE, PH, VN, BG, HR, MC",
     "为 High 等级公司提供多样的高风险但非全面制裁辖区",
     "scripts/generate-companies.ts （生成器）"),
    ("FSI hubs 字典", "金融保密辖区分档",
     "UAE 78, Monaco 70, Singapore 70, Switzerland 65, HK 62, ...",
     "Tax Justice Network FSI 排名手工分档",
     "Country Context Enrichment → Financial Secrecy"),
]

hs_headers = ["常量名", "中文名", "成员（ISO-2 国家代码）", "选择依据", "用于哪个因子"]
for i, h in enumerate(hs_headers, 1):
    c = hs.cell(row=3, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER

for i, row in enumerate(hardcoded, 1):
    r = 3 + i
    for col, val in enumerate(row, 1):
        cell = hs.cell(row=r, column=col, value=val)
        cell.alignment = WRAP
        cell.border = BORDER
        if col == 1:
            cell.font = Font(name="Courier New", size=10, bold=True)
        else:
            cell.font = Font(name=ARIAL, size=10)

for col, w in zip("ABCDE", [24, 22, 50, 40, 34]):
    hs.column_dimensions[col].width = w
for r in range(4, 4 + len(hardcoded)):
    hs.row_dimensions[r].height = 60

# Save
out_path = "/Users/cloud/Documents/aml/AML_Scoring_Spec.xlsx"
wb.save(out_path)
print(f"Wrote: {out_path}")
