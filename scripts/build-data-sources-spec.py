"""Generate AML_Data_Sources.xlsx — comprehensive catalog of 20 data sources
discussed in the AML Risk Watch project (12 currently integrated + 8 commercial
candidates evaluated for the dual-track proposal).

All column headers and category labels are in Chinese; source proper names
(FATF, OFAC, etc.) are kept in their original brand form.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F2937")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

FILL_INTEGRATED = PatternFill("solid", start_color="DCFCE7")  # 已集成 - 深绿
FILL_PUBLIC = PatternFill("solid", start_color="ECFDF5")       # 公开未集成 - 浅薄荷
FILL_COMMERCIAL = PatternFill("solid", start_color="EEF2FF")   # 商业候选 - 淡紫

TYPE_COLORS = {
    "API-first": "10B981",
    "Both": "0EA5E9",
    "List": "F59E0B",
}

# ----------------------------------------------------------------------------
# 20 数据源主表
# 字段: # / 名称 / 提供方 / 类型 / 数据格式 / 更新频率 / 数据总量 /
#       是否有 API / 免费/商业 / 是否集成 / 主要用途 / URL / 备注
# ----------------------------------------------------------------------------
SOURCES = [
    # ===== 已集成的公开源 =====
    (1, "FATF Lists (High-Risk & Monitored Jurisdictions)",
     "Financial Action Task Force",
     "List", "HTML + PDF 公开声明",
     "每年 3 次（2 月 / 6 月 / 10 月）",
     "约 25 个司法辖区（3 黑 + ~22 灰）",
     "无（需爬取）", "免费", "已集成",
     "国家级黑/灰名单",
     "https://www.fatf-gafi.org/en/publications/High-risk-and-other-monitored-jurisdictions.html",
     "权威国家级制裁分类"),
    (2, "OFAC SDN List",
     "U.S. Treasury OFAC",
     "Both", "XML / CSV / PDF + 搜索 API",
     "每周多次（部分情况每天）",
     "约 12,000 条（个人 + 实体）",
     "有（Trade.gov CSL API）", "免费", "已集成",
     "个人/实体制裁筛查",
     "https://sanctionssearch.ofac.treas.gov/",
     "美国主制裁名单；CSL API 整合多个美国名单"),
    (3, "OFAC Country Programs",
     "U.S. Treasury OFAC",
     "List", "HTML + PDF 网页文档",
     "不定期（行政令变更触发）",
     "约 30+ 国家级制裁项目",
     "无", "免费", "已集成",
     "国家级制裁项目",
     "https://ofac.treasury.gov/sanctions-programs-and-country-information",
     "全面制裁（古巴/伊朗/朝鲜/叙利亚）vs 部门制裁（俄/委/白）"),
    (4, "UN Consolidated Sanctions List",
     "United Nations Security Council",
     "List", "XML + HTML",
     "每月（安理会会后可能加更）",
     "约 1,000 条（个人 + 实体）",
     "无（仅下载）", "免费", "已集成",
     "联合国制裁",
     "https://www.un.org/securitycouncil/content/un-sc-consolidated-list",
     "UN 宪章第七章下成员国强制执行"),
    (5, "EU Consolidated Financial Sanctions List",
     "European Commission FSD",
     "Both", "XML",
     "每月",
     "约 3,000 条（个人 + 实体）",
     "有（CIRCABC API）", "免费", "已集成",
     "欧盟金融制裁筛查",
     "https://webgate.ec.europa.eu/fsd/fsf",
     "欧盟 CFSP 金融制裁主要来源"),
    (6, "Basel AML Index",
     "Basel Institute on Governance",
     "List", "PDF 报告 + Excel",
     "每年一次",
     "152 个国家评分",
     "仅订阅 API", "免费 / 部分付费", "已集成",
     "国家洗钱/恐融风险指数 (0-10)",
     "https://index.baselgovernance.org",
     "17 项指标合成；监管和银行界广泛引用"),
    (7, "Transparency International CPI",
     "Transparency International",
     "List", "HTML + Excel",
     "每年一次（1 月）",
     "180 个国家评分",
     "无（仅下载）", "免费", "已集成",
     "国家清廉指数",
     "https://www.transparency.org/en/cpi",
     "0=极腐败，100=极清廉。反向映射为风险分"),
    (8, "World Bank WGI",
     "World Bank Group",
     "Both", "CSV / Excel + API",
     "每年一次（9 月）",
     "215 个国家 × 6 个治理维度",
     "有（World Bank Indicators API）", "免费", "已集成",
     "国家治理指标",
     "https://info.worldbank.org/governance/wgi/",
     "六大维度；本项目使用 Control of Corruption (-2.5..2.5)"),
    (9, "WJP Rule of Law Index",
     "World Justice Project",
     "List", "PDF + Excel",
     "每年一次（10 月）",
     "142 个国家评分",
     "无公开 API", "免费", "已集成（代理）",
     "国家法治评分",
     "https://worldjusticeproject.org/rule-of-law-index",
     "8 大因子 / 44 子因子。本项目当前用 WGI 代理"),
    (10, "Tax Justice Network FSI",
     "Tax Justice Network",
     "List", "PDF 报告 + Excel",
     "每两年一次",
     "141 个司法辖区评分",
     "无", "免费", "已集成（代理）",
     "金融保密度",
     "https://taxjustice.net/topics/financial-secrecy-index/",
     "保密枢纽前列：美国、瑞士、新加坡、香港、卢森堡、阿联酋"),
    (11, "OpenSanctions",
     "OpenSanctions e.V.",
     "API-first", "JSON API + 批量 dump",
     "每天（免费）/ 每小时（付费）",
     "约 60 万条聚合记录（PEP ~25 万）",
     "有（REST + GraphQL）", "免费 / 部分付费", "已集成",
     "聚合制裁 + PEP",
     "https://www.opensanctions.org",
     "开源聚合器，覆盖全球主要监控名单"),
    (12, "Control Risks Country Risk",
     "Control Risks",
     "Both", "Web 平台 + 报告",
     "实时 + RiskMap 年度",
     "约 200 个国家 × 多项风险维度",
     "有（商业订阅）", "商业", "已集成（引用）",
     "地缘政治 / 运营风险",
     "https://www.controlrisks.com/our-thinking/insights/riskmap",
     "用于中转枢纽和规避因子的逻辑参考"),

    # ===== 商业候选源（未集成）=====
    (13, "Dow Jones Risk & Compliance / Factiva",
     "Dow Jones",
     "API-first", "JSON API",
     "每天（制裁）/ 24-48h（负面媒体）",
     "约 500 万条画像（PEP / 制裁 / 负面）",
     "有（DJ Risk API）", "商业", "未集成",
     "PEP、负面媒体、监管处罚",
     "https://www.dowjones.com/professional/risk/",
     "联想现有供应商；按席位约 10 万-200 万美元/年"),
    (14, "Refinitiv (LSEG) World-Check One",
     "London Stock Exchange Group",
     "API-first", "JSON API",
     "一天多次",
     "约 600+ 万条画像",
     "有（World-Check One API）", "商业", "未集成",
     "PEP、制裁、负面媒体",
     "https://www.refinitiv.com/en/products/world-check-kyc-screening",
     "最被引用的商业数据集；实体解析强"),
    (15, "Moody's Orbis + GRID",
     "Moody's Analytics",
     "Both", "JSON API + CSV 导出",
     "GRID 每天 / Orbis 每月",
     "Orbis 5 亿+ 公司；GRID 约 1300 万实体",
     "有（Orbis + GRID API）", "商业", "未集成",
     "公司数据 + 负面媒体组合",
     "https://www.moodys.com/web/en/us/capabilities/master-data/orbis.html",
     "公司层级 / UBO 深挖最强"),
    (16, "LexisNexis WorldCompliance + Nexis Diligence",
     "LexisNexis",
     "API-first", "JSON API",
     "每天",
     "约 300 万条画像 + 法律记录",
     "有（Nexis API）", "商业", "未集成",
     "法律 / 诉讼 / 合规",
     "https://risk.lexisnexis.com/global/en/products/worldcompliance-online-search-tool",
     "法律记录覆盖最强；实时 PEP 弱于 DJ"),
    (17, "ComplyAdvantage",
     "ComplyAdvantage",
     "API-first", "现代 REST/GraphQL API",
     "实时",
     "约 500 万条聚合数据",
     "有（AI 驱动 API）", "商业", "未集成",
     "AI 驱动 AML SaaS",
     "https://complyadvantage.com",
     "中型机构友好（约 1.5 万-5 万美元/年）；接入最快"),
    (18, "Sayari",
     "Sayari",
     "API-first", "图数据 API",
     "持续更新",
     "约 40 亿条全球记录（公司 + 关系）",
     "有（Sayari Graph API）", "商业", "未集成",
     "网络图谱 / 规避追溯",
     "https://sayari.com",
     "制裁规避 / 壳公司追溯一流"),
    (19, "Dun & Bradstreet (D&B)",
     "Dun & Bradstreet",
     "API-first", "Direct+ JSON API",
     "每天（信用 / 风险打分）",
     "5 亿+ 全球企业",
     "有（D&B Direct+）", "商业", "未集成",
     "公司信用 + 实体数据",
     "https://www.dnb.com/products/data/dnb-direct.html",
     "联想已用于中国 TPDD；CN/ROW 拆分对 API 不便"),
    (20, "GDELT",
     "The GDELT Project",
     "API-first", "DOC 2.0 API + BigQuery",
     "每 15 分钟",
     "25 亿+ 全球事件（自 1979 起）",
     "有（免费公共 API）", "免费", "未集成",
     "全球新闻事件监控",
     "https://www.gdeltproject.org/",
     "免费的负面媒体替代源；缺乏合规分类标签"),
]

CATEGORY_OF_ROW = {n: ("public" if n <= 12 else "commercial") for n in range(1, 21)}

# ============================================================
# Build workbook
# ============================================================
wb = Workbook()

# ---- Sheet 1: 主表 ----
ws = wb.active
ws.title = "全部数据源"

ws["A1"] = "AML 数据源完整目录"
ws["A1"].font = Font(name=ARIAL, bold=True, size=16, color="111827")
ws.merge_cells("A1:M1")

ws["A2"] = "已集成 12 条 + 双轨方案中评估的 8 条商业候选 = 20 条"
ws["A2"].font = Font(name=ARIAL, italic=True, size=10, color="6B7280")
ws.merge_cells("A2:M2")

headers = ["#", "来源名称", "提供方", "类型", "数据格式", "更新频率",
           "数据总量", "是否有 API", "免费/商业", "是否集成",
           "主要用途", "URL", "备注"]
HEADER_ROW = 4
for i, h in enumerate(headers, 1):
    c = ws.cell(row=HEADER_ROW, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = WRAP_CENTER
    c.border = BORDER

ws.row_dimensions[HEADER_ROW].height = 36

# Data rows
for r_idx, row in enumerate(SOURCES, start=HEADER_ROW + 1):
    (n, name, provider, type_, fmt, cadence, volume, api, pricing, integrated,
     use, url, notes) = row
    cat = CATEGORY_OF_ROW[n]
    base_fill = FILL_INTEGRATED if integrated.startswith("已集成") else (
        FILL_PUBLIC if cat == "public" else FILL_COMMERCIAL)

    cells = [n, name, provider, type_, fmt, cadence, volume, api, pricing,
             integrated, use, url, notes]
    for c_idx, val in enumerate(cells, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = base_fill
        cell.border = BORDER
        cell.alignment = WRAP
        cell.font = Font(name=ARIAL, size=10)

    # # column
    ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=r_idx, column=1).font = Font(name=ARIAL, size=10, bold=True, color="64748B")
    # 类型
    type_cell = ws.cell(row=r_idx, column=4)
    color = TYPE_COLORS.get(type_, "64748B")
    type_cell.font = Font(name=ARIAL, size=10, bold=True, color=color)
    type_cell.alignment = WRAP_CENTER
    # 数据总量 (mono, bold for emphasis)
    vol_cell = ws.cell(row=r_idx, column=7)
    vol_cell.font = Font(name="Courier New", size=10, bold=True, color="111827")
    vol_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    # API?
    api_cell = ws.cell(row=r_idx, column=8)
    if api.startswith("有"):
        api_cell.font = Font(name=ARIAL, size=10, bold=True, color="10B981")
    elif api.startswith("仅订阅"):
        api_cell.font = Font(name=ARIAL, size=10, bold=True, color="F59E0B")
    else:
        api_cell.font = Font(name=ARIAL, size=10, color="EF4444")
    api_cell.alignment = WRAP_CENTER
    # 免费/商业
    pricing_cell = ws.cell(row=r_idx, column=9)
    pricing_cell.alignment = WRAP_CENTER
    if pricing.startswith("免费"):
        pricing_cell.font = Font(name=ARIAL, size=10, bold=True, color="10B981")
    elif pricing.startswith("商业"):
        pricing_cell.font = Font(name=ARIAL, size=10, bold=True, color="7C3AED")
    # 是否集成
    int_cell = ws.cell(row=r_idx, column=10)
    int_cell.alignment = WRAP_CENTER
    if integrated.startswith("已集成"):
        int_cell.font = Font(name=ARIAL, size=10, bold=True, color="10B981")
    else:
        int_cell.font = Font(name=ARIAL, size=10, color="6B7280")
    # URL
    url_cell = ws.cell(row=r_idx, column=12)
    url_cell.hyperlink = url
    url_cell.font = Font(name=ARIAL, size=9, color="0563C1", underline="single")

# Column widths
widths = [4, 32, 28, 11, 22, 24, 26, 22, 14, 14, 24, 36, 46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "C5"
for r_idx in range(HEADER_ROW + 1, HEADER_ROW + 1 + len(SOURCES)):
    ws.row_dimensions[r_idx].height = 64

# ---- Sheet 2: 按类型分布 ----
ws2 = wb.create_sheet("按类型分布")
ws2["A1"] = "按数据源类型分布"
ws2["A1"].font = Font(name=ARIAL, bold=True, size=14)
ws2.merge_cells("A1:E1")

for i, h in enumerate(["类型", "数量", "占比", "来源"], 1):
    c = ws2.cell(row=3, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = WRAP_CENTER
    c.border = BORDER

type_groups = {}
for s in SOURCES:
    t = s[3]
    type_groups.setdefault(t, []).append(s[1])

row_no = 4
for t in ["List", "API-first", "Both"]:
    items = type_groups.get(t, [])
    ws2.cell(row=row_no, column=1, value=t).font = Font(name=ARIAL, bold=True, size=11, color=TYPE_COLORS.get(t, "111827"))
    ws2.cell(row=row_no, column=2, value=len(items)).alignment = WRAP_CENTER
    ws2.cell(row=row_no, column=3, value=f"=B{row_no}/SUM($B$4:$B$6)").number_format = "0.0%"
    ws2.cell(row=row_no, column=3).alignment = WRAP_CENTER
    ws2.cell(row=row_no, column=4, value=" · ".join(items)).alignment = WRAP
    ws2.cell(row=row_no, column=4).font = Font(name=ARIAL, size=9)
    for col in range(1, 5):
        ws2.cell(row=row_no, column=col).border = BORDER
    row_no += 1

ws2.cell(row=row_no, column=1, value="合计").font = Font(name=ARIAL, bold=True, size=11)
ws2.cell(row=row_no, column=2, value=f"=SUM(B4:B{row_no-1})").font = Font(name=ARIAL, bold=True, size=11)
ws2.cell(row=row_no, column=2).alignment = WRAP_CENTER
ws2.cell(row=row_no, column=3, value=f"=SUM(C4:C{row_no-1})").number_format = "0.0%"
ws2.cell(row=row_no, column=3).alignment = WRAP_CENTER
ws2.cell(row=row_no, column=3).font = Font(name=ARIAL, bold=True, size=11)
for col in range(1, 5):
    ws2.cell(row=row_no, column=col).border = BORDER

def_row = row_no + 3
ws2.cell(row=def_row, column=1, value="类型说明").font = Font(name=ARIAL, bold=True, size=12)
defs = [
    ("List", "静态发布物，需手动下载（CSV/XML/PDF/Excel）。刷新 = 重新下载。", "F59E0B"),
    ("API-first", "主要通过 HTTP API 查询。最适合自动化集成。", "10B981"),
    ("Both", "既提供 List 下载，又提供可查询的 API。", "0EA5E9"),
]
for i, (t, d, color) in enumerate(defs):
    r = def_row + 1 + i
    ws2.cell(row=r, column=1, value=t).font = Font(name=ARIAL, bold=True, size=11, color=color)
    ws2.cell(row=r, column=2, value=d).alignment = WRAP
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 90
for r in range(4, 7):
    ws2.row_dimensions[r].height = 80

# ---- Sheet 3: 按更新频率分布 ----
ws3 = wb.create_sheet("按更新频率分布")
ws3["A1"] = "按更新频率分布"
ws3["A1"].font = Font(name=ARIAL, bold=True, size=14)
ws3.merge_cells("A1:E1")

for i, h in enumerate(["更新频率", "数量", "占比", "来源", "运营含义"], 1):
    c = ws3.cell(row=3, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = WRAP_CENTER
    c.border = BORDER

cadence_groups = [
    ("实时 / 每 15 分钟",
     ["OpenSanctions", "ComplyAdvantage", "GDELT"],
     "可用于交易实时拦截。必须 API 集成。"),
    ("每天",
     ["OFAC SDN List", "Dow Jones Risk & Compliance / Factiva",
      "Refinitiv (LSEG) World-Check One",
      "LexisNexis WorldCompliance + Nexis Diligence",
      "Dun & Bradstreet (D&B)", "Moody's Orbis + GRID"],
     "对新客户次日批量筛查足够。"),
    ("每月",
     ["UN Consolidated Sanctions List", "EU Consolidated Financial Sanctions List",
      "Moody's Orbis + GRID"],
     "可用于定期重筛；可能漏掉月内新增。"),
    ("每年 / 每两年",
     ["FATF Lists（每年 3 次）", "Basel AML Index",
      "Transparency International CPI", "World Bank WGI",
      "WJP Rule of Law Index", "Tax Justice Network FSI"],
     "国家级指数 — 发布时刷新 + 季度抽查就足够。"),
    ("不定期",
     ["OFAC Country Programs"],
     "由行政令驱动。订阅 OFAC 公告订阅。"),
]
row_no = 4
for cadence, items, impl in cadence_groups:
    ws3.cell(row=row_no, column=1, value=cadence).font = Font(name=ARIAL, bold=True, size=11)
    ws3.cell(row=row_no, column=2, value=len(items)).alignment = WRAP_CENTER
    ws3.cell(row=row_no, column=3, value=f"=B{row_no}/SUM($B$4:$B${4+len(cadence_groups)-1})").number_format = "0.0%"
    ws3.cell(row=row_no, column=3).alignment = WRAP_CENTER
    ws3.cell(row=row_no, column=4, value=" · ".join(items)).alignment = WRAP
    ws3.cell(row=row_no, column=4).font = Font(name=ARIAL, size=9)
    ws3.cell(row=row_no, column=5, value=impl).alignment = WRAP
    ws3.cell(row=row_no, column=5).font = Font(name=ARIAL, italic=True, size=10, color="475569")
    for col in range(1, 6):
        ws3.cell(row=row_no, column=col).border = BORDER
    ws3.row_dimensions[row_no].height = 64
    row_no += 1

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 8
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 80
ws3.column_dimensions["E"].width = 60

# ---- Sheet 4: 集成难度 ----
ws4 = wb.create_sheet("集成难度")
ws4["A1"] = "按集成难度分类"
ws4["A1"].font = Font(name=ARIAL, bold=True, size=14)
ws4.merge_cells("A1:D1")

for i, h in enumerate(["难度", "工作量", "来源", "集成方式"], 1):
    c = ws4.cell(row=3, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = WRAP_CENTER
    c.border = BORDER

automation = [
    ("🟢 容易", "数小时即可",
     ["OpenSanctions", "GDELT", "World Bank WGI",
      "OFAC SDN List（通过 Trade.gov CSL API）",
      "Dow Jones Risk & Compliance", "Refinitiv World-Check One",
      "LexisNexis WorldCompliance", "ComplyAdvantage",
      "Sayari", "Dun & Bradstreet", "Moody's Orbis + GRID"],
     "直接 API 调用。JSON 进 JSON 出。加缓存层 + 重试即可。"),
    ("🟡 中等", "数天",
     ["UN Consolidated", "EU Consolidated",
      "Transparency International CPI", "WJP Rule of Law Index",
      "Tax Justice Network FSI"],
     "下载结构化文件（XML/Excel）+ 解析 + 写入本地 JSON。"),
    ("🔴 手动", "数周（或接受手动更新）",
     ["FATF Lists", "OFAC Country Programs", "Basel AML Index"],
     "PDF / HTML 解析或人工录入。这些源更新频率低，季度刷新通常足够。"),
]
row_no = 4
for tier, effort, items, how in automation:
    ws4.cell(row=row_no, column=1, value=tier).font = Font(name=ARIAL, bold=True, size=12)
    ws4.cell(row=row_no, column=2, value=effort).alignment = WRAP
    ws4.cell(row=row_no, column=2).font = Font(name=ARIAL, italic=True, size=10)
    ws4.cell(row=row_no, column=3, value=" · ".join(items)).alignment = WRAP
    ws4.cell(row=row_no, column=3).font = Font(name=ARIAL, size=9)
    ws4.cell(row=row_no, column=4, value=how).alignment = WRAP
    ws4.cell(row=row_no, column=4).font = Font(name=ARIAL, size=10)
    for col in range(1, 5):
        ws4.cell(row=row_no, column=col).border = BORDER
    ws4.row_dimensions[row_no].height = 80
    row_no += 1

ws4.column_dimensions["A"].width = 14
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 70
ws4.column_dimensions["D"].width = 50

# ---- Sheet 5: 实务建议 ----
ws5 = wb.create_sheet("实务建议")
ws5["A1"] = "给 ECO 的实务建议"
ws5["A1"].font = Font(name=ARIAL, bold=True, size=14)
ws5.merge_cells("A1:C1")

recs = [
    ("制裁筛查（OFAC / UN / EU）",
     "三个都既有 List 又有 API。建议直接用 Trade.gov 综合筛查名单 API 把 OFAC + 其他美国名单聚合成一次调用。",
     "公开覆盖与 DJ 同等 — 此场景商业版无优势。"),
    ("国家级指数（Basel / CPI / WGI / WJP / FSI）",
     "每年/每两年更新。每年发布时重新下载源文件并更新 data/countries.json 即可。",
     "List 模式足够，无需 API。当前 dashboard 就在 data/countries.json 中。"),
    ("负面媒体（公开源最大短板）",
     "免费版：GDELT（每 15 分钟，无合规分类标签）；商业版：DJ / Refinitiv / LexisNexis / ComplyAdvantage 都有分析师打标的合规分类。",
     "DJ 订阅最值钱的地方。双轨方案保留 DJ 的最强理由。"),
    ("PEP 覆盖",
     "公开版：OpenSanctions（约 25 万 PEP，周更）；商业版：DJ / Refinitiv（约 300 万 PEP，日更，含地市级官员）。",
     "新兴市场辖区差距明显；G7 国家差距小。"),
    ("受益所有权（UBO）数据",
     "公开版：欧盟/英国之外覆盖薄弱；商业版：Sayari / Moody's Orbis / D&B 覆盖明显更深。",
     "如果纳入规避追溯需求，Sayari 是最强候选。"),
    ("实体消歧",
     "商业供应商已预解析到 LEI / 唯一 ID；公开源需自建名称匹配管道。",
     "保留至少一家商业供应商的运营效率论据。"),
]

for i, h in enumerate(["需求", "可用来源", "建议"], 1):
    c = ws5.cell(row=3, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = WRAP_CENTER
    c.border = BORDER

row_no = 4
for need, avail, rec in recs:
    ws5.cell(row=row_no, column=1, value=need).font = Font(name=ARIAL, bold=True, size=11, color="1F2937")
    ws5.cell(row=row_no, column=1).alignment = WRAP
    ws5.cell(row=row_no, column=2, value=avail).alignment = WRAP
    ws5.cell(row=row_no, column=2).font = Font(name=ARIAL, size=10)
    ws5.cell(row=row_no, column=3, value=rec).alignment = WRAP
    ws5.cell(row=row_no, column=3).font = Font(name=ARIAL, italic=True, size=10, color="475569")
    for col in range(1, 4):
        ws5.cell(row=row_no, column=col).border = BORDER
    ws5.row_dimensions[row_no].height = 80
    row_no += 1

ws5.column_dimensions["A"].width = 28
ws5.column_dimensions["B"].width = 65
ws5.column_dimensions["C"].width = 55

# Save
out = "/Users/cloud/Documents/aml/AML_Data_Sources.xlsx"
wb.save(out)
print(f"Wrote: {out}")
