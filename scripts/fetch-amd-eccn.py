"""Download AMD's public HTS/ECCN master PDF and convert it to:

  - data/amd-eccn.json           structured records (for the /exports page)
  - data/amd-eccn-meta.json      stats (for overview cards)
  - public/amd-eccn.xlsx         downloadable Excel workbook

Source:
  https://www.amd.com/en/legal/compliance/trade-compliance.html
  → "AMD Product HTS/ECCN List" → product-master.pdf
"""
from __future__ import annotations

import json
import re
import sys
import urllib.request
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PDF_URL = "https://www.amd.com/content/dam/amd/en/documents/legal/product-master.pdf"
SOURCE_URL = "https://www.amd.com/en/legal/compliance/trade-compliance.html"

ROOT = Path(__file__).resolve().parent.parent
PDF_CACHE = ROOT / "scripts" / ".cache" / "amd-product-master.pdf"
DATA_JSON = ROOT / "data" / "amd-eccn.json"
META_JSON = ROOT / "data" / "amd-eccn-meta.json"
XLSX_PATH = ROOT / "public" / "amd-eccn.xlsx"

HEADER_RE = re.compile(
    r"^\s*Product\s+Number\s+US\s+ECCN\s+US\s+HS\s+CCATS\s+Meets\s+3A090",
    re.IGNORECASE,
)
# Example rows (CCATS can be e.g. "G177385", "Self", "Self / B.1", "-", "N/A"):
#   "100-000000009   5A992.c    8542310045 G177385       N"
#   "100-000001440   5A992.c    8542310045 Self / B.1    N"
#   "100-438153      EAR99      8473301180 Self          N"
ROW_RE = re.compile(
    r"^(?P<part>[A-Za-z0-9][A-Za-z0-9\-./]+)\s+"
    r"(?P<eccn>[A-Za-z0-9][A-Za-z0-9.\-]*)\s+"
    r"(?P<hs>\d{8,12})\s+"
    r"(?P<ccats>.+?)\s+"
    r"(?P<meets>[YN])\s*$"
)


def download_pdf() -> Path:
    PDF_CACHE.parent.mkdir(parents=True, exist_ok=True)
    if PDF_CACHE.exists() and PDF_CACHE.stat().st_size > 10000:
        print(f"  Using cached PDF: {PDF_CACHE}")
        return PDF_CACHE
    req = urllib.request.Request(
        PDF_URL,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
            ),
            "Referer": SOURCE_URL,
            "Accept": "application/pdf,*/*",
        },
    )
    with urllib.request.urlopen(req, timeout=60) as r, PDF_CACHE.open("wb") as f:
        f.write(r.read())
    print(f"  Downloaded {PDF_CACHE} ({PDF_CACHE.stat().st_size // 1024} KB)")
    return PDF_CACHE


def parse_pdf(pdf_path: Path) -> tuple[list[dict], str | None]:
    """Extract all product rows from the PDF. Returns (rows, classification_date)."""
    rows: list[dict] = []
    seen_part_numbers: set[str] = set()
    classification_date: str | None = None

    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                # Capture the classification date from page 1
                m = re.search(r"Classification Data as of\s+(.+)$", line, re.I)
                if m and not classification_date:
                    classification_date = m.group(1).strip().rstrip(".")
                if HEADER_RE.match(line):
                    continue
                match = ROW_RE.match(line)
                if not match:
                    continue
                part = match.group("part")
                if part in seen_part_numbers:
                    continue
                seen_part_numbers.add(part)
                rows.append(
                    {
                        "part_number": part,
                        "us_eccn": match.group("eccn"),
                        "us_hs": match.group("hs"),
                        "ccats": match.group("ccats"),
                        "meets_3A090_a_1": match.group("meets"),
                    }
                )
            if (page_idx + 1) % 20 == 0:
                print(f"  Parsed page {page_idx + 1} / {len(pdf.pages)} — {len(rows)} rows so far")

    print(f"  Total rows: {len(rows)} (classification date: {classification_date})")
    return rows, classification_date


def write_xlsx(rows: list[dict], classification_date: str | None) -> None:
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    header_fill = PatternFill("solid", start_color="1F2937")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Sheet 1: About
    info = wb.active
    info.title = "About"
    info["A1"] = "AMD Product HTS/ECCN List"
    info["A1"].font = Font(name="Arial", bold=True, size=16)
    info.merge_cells("A1:C1")
    info["A3"] = "Source page"
    info["B3"] = SOURCE_URL
    info["B3"].hyperlink = SOURCE_URL
    info["B3"].font = Font(name="Arial", color="0563C1", underline="single")
    info["A4"] = "Upstream PDF"
    info["B4"] = PDF_URL
    info["B4"].hyperlink = PDF_URL
    info["B4"].font = Font(name="Arial", color="0563C1", underline="single")
    info["A5"] = "Snapshot fetched (UTC)"
    info["B5"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    if classification_date:
        info["A6"] = "Classification as of"
        info["B6"] = classification_date
        info["B6"].font = Font(name="Arial", bold=True)
    info["A7"] = "Records"
    info["B7"] = len(rows)
    info["B7"].font = Font(name="Arial", bold=True)
    info["A9"] = "Field glossary"
    info["A9"].font = Font(name="Arial", bold=True, size=12)
    glossary = [
        ("part_number", "AMD product number"),
        ("us_eccn", "U.S. Export Control Classification Number (EAR)"),
        ("us_hs", "U.S. Harmonized System code used for customs"),
        ("ccats", "Commodity Classification Automated Tracking System reference (BIS)"),
        ("meets_3A090_a_1", "Whether the part meets ECCN 3A090.a.1 AI-chip performance threshold (Y/N)"),
    ]
    for i, (field, desc) in enumerate(glossary, start=11):
        info.cell(row=i, column=1, value=field).font = Font(name="Courier New", size=10)
        info.cell(row=i, column=2, value=desc).alignment = wrap
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 70

    # Sheet 2: All Parts
    ws = wb.create_sheet("All Parts")
    columns = [
        ("part_number", "Part Number", 26),
        ("us_eccn", "US ECCN", 14),
        ("us_hs", "US HS", 14),
        ("ccats", "CCATS", 14),
        ("meets_3A090_a_1", "Meets 3A090.a.1", 16),
    ]
    for i, (_, label, w) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows, start=2):
        for ci, (key, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=ri, column=ci, value=row.get(key, ""))
            cell.border = border
            if ci in (1, 2, 3, 4):
                cell.font = Font(name="Courier New", size=10)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    # Sheet 3: ECCN Distribution
    dist = Counter(r["us_eccn"] for r in rows)
    ws2 = wb.create_sheet("ECCN Distribution")
    for i, h in enumerate(["ECCN", "Count", "% of total"], start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    for i, (eccn, n) in enumerate(dist.most_common(), start=2):
        ws2.cell(row=i, column=1, value=eccn).font = Font(name="Courier New", size=10)
        ws2.cell(row=i, column=2, value=n).alignment = Alignment(horizontal="right")
        ws2.cell(row=i, column=3, value=f"=B{i}/B${len(dist)+2}").number_format = "0.0%"
        for c in range(1, 4):
            ws2.cell(row=i, column=c).border = border
    total_r = len(dist) + 2
    ws2.cell(row=total_r, column=1, value="Total").font = Font(name="Arial", bold=True)
    ws2.cell(row=total_r, column=2, value=f"=SUM(B2:B{total_r-1})").font = Font(name="Arial", bold=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14
    ws2.freeze_panes = "A2"

    wb.save(XLSX_PATH)


def write_meta(rows: list[dict], classification_date: str | None) -> None:
    eccn_counts = Counter(r["us_eccn"] for r in rows)
    hs_counts = Counter(r["us_hs"] for r in rows)
    ccats_counts = Counter(r["ccats"] for r in rows)
    meets_counts = Counter(r["meets_3A090_a_1"] for r in rows)
    meta = {
        "total": len(rows),
        "fetchedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "classificationDate": classification_date,
        "sourcePage": SOURCE_URL,
        "pdfUrl": PDF_URL,
        "distinctEccn": len(eccn_counts),
        "distinctHs": len(hs_counts),
        "distinctCcats": len(ccats_counts),
        "topEccn": [{"eccn": k, "count": v} for k, v in eccn_counts.most_common(10)],
        "topHs": [{"hs": k, "count": v} for k, v in hs_counts.most_common(10)],
        "meets3A090": [{"meets": k, "count": v} for k, v in meets_counts.most_common()],
    }
    META_JSON.write_text(json.dumps(meta, indent=2, ensure_ascii=False))


def main() -> int:
    print(f"Fetching {PDF_URL} …")
    pdf_path = download_pdf()
    print("Parsing PDF …")
    rows, classification_date = parse_pdf(pdf_path)
    if not rows:
        print("No rows parsed — regex may be off.", file=sys.stderr)
        return 1

    DATA_JSON.parent.mkdir(parents=True, exist_ok=True)
    DATA_JSON.write_text(json.dumps(rows, ensure_ascii=False))
    print(f"  Wrote {DATA_JSON.relative_to(ROOT)} ({DATA_JSON.stat().st_size // 1024} KB)")

    write_meta(rows, classification_date)
    print(f"  Wrote {META_JSON.relative_to(ROOT)}")

    write_xlsx(rows, classification_date)
    print(f"  Wrote {XLSX_PATH.relative_to(ROOT)} ({XLSX_PATH.stat().st_size // 1024} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
