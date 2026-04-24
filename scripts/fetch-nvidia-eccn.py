"""Fetch NVIDIA ECCN parts classification from their public API and emit:
  - data/nvidia-eccn.json           (for the /exports page to render)
  - public/nvidia-eccn.xlsx         (for user download)
  - data/nvidia-eccn-meta.json      (stats for page overview)

Source URL:
  https://api-prod.nvidia.com/services/eccn/v1/getECCN
  (discovered inside /content/dam/en-zz/Solutions/eccn-portal/eccn.js which is
   referenced on https://www.nvidia.com/en-us/about-nvidia/company-policies/export-regulations/)
"""
from __future__ import annotations

import json
import sys
import urllib.request
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

API_URL = "https://api-prod.nvidia.com/services/eccn/v1/getECCN"
ROOT = Path(__file__).resolve().parent.parent
DATA_JSON = ROOT / "data" / "nvidia-eccn.json"
META_JSON = ROOT / "data" / "nvidia-eccn-meta.json"
XLSX_PATH = ROOT / "public" / "nvidia-eccn.xlsx"
SOURCE_URL = "https://www.nvidia.com/en-us/about-nvidia/company-policies/export-regulations/"


def fetch() -> list[dict]:
    req = urllib.request.Request(
        API_URL,
        headers={
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Origin": "https://www.nvidia.com",
            "Referer": "https://www.nvidia.com/",
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
            ),
        },
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r)


def write_xlsx(rows: list[dict]) -> None:
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    header_fill = PatternFill("solid", start_color="1F2937")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Sheet 1: About
    info = wb.active
    info.title = "About"
    info["A1"] = "NVIDIA Export Regulation Compliance — Parts Classification"
    info["A1"].font = Font(name="Arial", bold=True, size=16)
    info.merge_cells("A1:C1")
    info["A3"] = "Source page"
    info["B3"] = SOURCE_URL
    info["B3"].hyperlink = SOURCE_URL
    info["B3"].font = Font(name="Arial", color="0563C1", underline="single")
    info["A4"] = "Upstream API"
    info["B4"] = API_URL
    info["B4"].hyperlink = API_URL
    info["B4"].font = Font(name="Arial", color="0563C1", underline="single")
    info["A5"] = "Snapshot time (UTC)"
    info["B5"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    info["A6"] = "Records"
    info["B6"] = len(rows)
    info["B6"].font = Font(name="Arial", bold=True)
    info["A8"] = "Field glossary"
    info["A8"].font = Font(name="Arial", bold=True, size=12)
    glossary = [
        ("part_number", "NVIDIA part number"),
        ("part_description", "Human-readable description of the part"),
        ("part_type", "Internal NVIDIA part type"),
        ("tpp", "Total Processing Power (relevant for AI accelerator export rules)"),
        ("nv_hts", "Harmonized Tariff Schedule (HTS) classification used for customs"),
        ("nveccn", "Export Control Classification Number (US EAR)"),
        ("state", "Lifecycle state (PRODUCTION / EOL / …)"),
        ("mx_legacy_Material", "Legacy material code (optional)"),
    ]
    for i, (field, desc) in enumerate(glossary, start=10):
        info.cell(row=i, column=1, value=field).font = Font(name="Courier New", size=10)
        info.cell(row=i, column=2, value=desc).alignment = wrap
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 70

    # Sheet 2: All Parts
    ws = wb.create_sheet("All Parts")
    columns = [
        ("id", "ID", 8),
        ("part_number", "Part Number", 22),
        ("part_description", "Description", 45),
        ("part_type", "Type", 14),
        ("tpp", "TPP", 10),
        ("nv_hts", "HTS Code", 12),
        ("nveccn", "ECCN", 14),
        ("state", "State", 14),
        ("mx_legacy_Material", "Legacy Material", 16),
    ]
    for i, (_, label, w) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    for ri, row in enumerate(rows, start=2):
        for ci, (key, _, _) in enumerate(columns, start=1):
            v = row.get(key)
            if v in (None, "NULL"):
                v = ""
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            if ci == 2:
                cell.font = Font(name="Courier New", size=10)
            elif ci == 3:
                cell.alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    # Sheet 3: ECCN Distribution
    dist_ws = wb.create_sheet("ECCN Distribution")
    dist_ws["A1"] = "ECCN"
    dist_ws["B1"] = "Count"
    dist_ws["C1"] = "% of total"
    for col in range(1, 4):
        c = dist_ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    counts = Counter((r.get("nveccn") or "N/A") for r in rows)
    for i, (eccn, n) in enumerate(counts.most_common(), start=2):
        dist_ws.cell(row=i, column=1, value=eccn).font = Font(name="Courier New", size=10)
        dist_ws.cell(row=i, column=2, value=n).alignment = Alignment(horizontal="right")
        dist_ws.cell(row=i, column=3, value=f"=B{i}/B${len(counts)+2}").number_format = "0.0%"
        for col in range(1, 4):
            dist_ws.cell(row=i, column=col).border = border
    total_row = len(counts) + 2
    dist_ws.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", bold=True)
    dist_ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})").font = Font(name="Arial", bold=True)
    dist_ws.column_dimensions["A"].width = 18
    dist_ws.column_dimensions["B"].width = 10
    dist_ws.column_dimensions["C"].width = 12
    dist_ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)


def write_meta(rows: list[dict]) -> None:
    eccn_counts = Counter((r.get("nveccn") or "N/A") for r in rows)
    hts_counts = Counter((r.get("nv_hts") or "N/A") for r in rows)
    state_counts = Counter((r.get("state") or "N/A") for r in rows)
    type_counts = Counter((r.get("part_type") or "N/A") for r in rows)
    meta = {
        "total": len(rows),
        "fetchedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "sourcePage": SOURCE_URL,
        "apiUrl": API_URL,
        "distinctEccn": len(eccn_counts),
        "distinctHts": len(hts_counts),
        "topEccn": [{"eccn": k, "count": v} for k, v in eccn_counts.most_common(10)],
        "topHts": [{"hts": k, "count": v} for k, v in hts_counts.most_common(10)],
        "states": [{"state": k, "count": v} for k, v in state_counts.most_common()],
        "partTypes": [{"type": k, "count": v} for k, v in type_counts.most_common(10)],
    }
    META_JSON.write_text(json.dumps(meta, indent=2, ensure_ascii=False))


def main() -> int:
    print(f"Fetching {API_URL} …")
    rows = fetch()
    if not isinstance(rows, list):
        print(f"Unexpected payload type: {type(rows).__name__}", file=sys.stderr)
        return 1
    print(f"  Received {len(rows)} records")

    DATA_JSON.parent.mkdir(parents=True, exist_ok=True)
    DATA_JSON.write_text(json.dumps(rows, ensure_ascii=False))
    print(f"  Wrote {DATA_JSON.relative_to(ROOT)} ({DATA_JSON.stat().st_size // 1024} KB)")

    write_meta(rows)
    print(f"  Wrote {META_JSON.relative_to(ROOT)}")

    write_xlsx(rows)
    print(f"  Wrote {XLSX_PATH.relative_to(ROOT)} ({XLSX_PATH.stat().st_size // 1024} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
